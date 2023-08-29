from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from subprocess import CREATE_NO_WINDOW
from time import sleep
from PySimpleGUI import PySimpleGUI as sg
import openpyxl

#GUI
sg.theme('Reddit')
layout = [
    [sg.Text('Documento'), sg.Input(key='documento', do_not_clear=False)],
    [sg.Button('Pesquisar')],
    [sg.Text('', key='status')]
]
janela = sg.Window('Pesquisa por Documento', layout)

def search(expand):
    expand.click()
    proccess_id = driver.find_element(By.XPATH, "//span[@id='numeroProcesso']")
    proccess_id = proccess_id.text
    proccess_class = driver.find_element(By.XPATH, "//span[@id='classeProcesso']")
    proccess_class = proccess_class.text
    proccess_subject = driver.find_element(By.XPATH, "//span[@id='assuntoProcesso']")
    proccess_subject = proccess_subject.text
    proccess_foro = driver.find_element(By.XPATH, "//span[@id='foroProcesso']")
    proccess_foro = proccess_foro.text
    proccess_vara = driver.find_element(By.XPATH, "//span[@id='varaProcesso']")
    proccess_vara = proccess_vara.text
    proccess_judge = driver.find_element(By.XPATH, "//span[@id='juizProcesso']")
    proccess_judge = proccess_judge.text
    proccess_distruicao = driver.find_element(By.XPATH, "//div[@id='dataHoraDistribuicaoProcesso']")
    proccess_distruicao = proccess_distruicao.text
    proccess_controle = driver.find_element(By.XPATH, "//div[@id='numeroControleProcesso']")
    proccess_controle = proccess_controle.text
    proccess_area = driver.find_element(By.XPATH, "//div[@id='areaProcesso']")
    proccess_area = proccess_area.text
    proccess_valor = driver.find_element(By.XPATH, "//div[@id='valorAcaoProcesso']")
    proccess_valor = proccess_valor.text

    #movimentações
    data_movimentacoes = driver.find_elements(By.XPATH, "//tr[contains(@class, 'containerMovimentacao')]//td[contains(@class, 'dataMovimentacao')]")
    desc_movimentacoes = driver.find_elements(By.XPATH, "//tr[contains(@class, 'containerMovimentacao')]//td[contains(@class, 'descricaoMovimentacao')]")
    lista_data_movimentacoes = []
    lista_desc_movimentacoes = []
    for i in range(4):
        lista_data_movimentacoes.append(data_movimentacoes[i].text)
        lista_desc_movimentacoes.append(desc_movimentacoes[i].text)

    #criar a planilha
    workbook = openpyxl.load_workbook('dados.xlsx')

    if proccess_id in workbook.sheetnames:
        workbook_page = workbook[proccess_id]
    else:
        workbook_page = workbook.create_sheet(proccess_id)

    workbook_page['A1'].value = "Número do Processo"
    workbook_page['A2'].value = proccess_id
    workbook_page['B1'].value = "Classe do Processo"
    workbook_page['B2'].value = proccess_class
    workbook_page['C1'].value = "Assunto do Processo"
    workbook_page['C2'].value = proccess_subject
    workbook_page['D1'].value = "Foro do Processo"
    workbook_page['D2'].value = proccess_foro
    workbook_page['E1'].value = "Vara do Processo"
    workbook_page['E2'].value = proccess_vara
    workbook_page['F1'].value = "Juiz do Processo"
    workbook_page['F2'].value = proccess_judge
    workbook_page['G1'].value = "Distribuição do Processo"
    workbook_page['G2'].value = proccess_distruicao
    workbook_page['H1'].value = "Controle do Processo"
    workbook_page['H2'].value = proccess_controle
    workbook_page['I1'].value = "Área do Processo"
    workbook_page['I2'].value = proccess_area
    workbook_page['J1'].value = "Valor do Processo"
    workbook_page['J2'].value = proccess_valor
    workbook_page['K1'].value = "Data Movimentação"
    workbook_page['L1'].value = "Descrição Movimentação"

    for index, row in enumerate(workbook_page.iter_rows(min_row=2, max_row=len(lista_data_movimentacoes), min_col=11, max_col=11)):
        for cell in row:
            cell.value = lista_data_movimentacoes[index]

    for index, row in enumerate(workbook_page.iter_rows(min_row=2, max_row=len(lista_desc_movimentacoes), min_col=12, max_col=12)):
        for cell in row:
            cell.value = lista_desc_movimentacoes[index]

    workbook.save('dados.xlsx')
    driver.close()
    janela['status'].update('')
    sg.popup('Dados extraídos com sucesso, verifique a planilha!')

while True:
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_service = ChromeService()
    chrome_service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Chrome(service=chrome_service,options=chrome_options)
    eventos, valores = janela.read()
    if eventos == sg.WINDOW_CLOSED:
        if(driver):
            driver.close()
        break
    if eventos == 'Pesquisar':
        janela['status'].update('Buscando Dados...')
        if valores['documento'] != '':
            driver.get('https://esaj.tjsp.jus.br/cpopg/open.do')
            client_document = valores['documento']
            input_document = driver.find_element(By.XPATH, "//input[@id='campo_DOCPARTE']")
            dropdown_form = driver.find_element(By.XPATH, "//select[@id='cbPesquisa']")
            options_select = Select(dropdown_form)
            options_select.select_by_value('DOCPARTE')
            input_document.clear()
            input_document.send_keys(client_document)
            button_submit = driver.find_element(By.XPATH, "//input[@id='botaoConsultarProcessos']")
            button_submit.click()
            sleep(2)
            try:
                expand = driver.find_element(By.XPATH, "//a[@href='#maisDetalhes']")
                search(expand)
            except Exception as error:
                sg.popup_error('Não encontrou dados com o Documento informado.')
                valores['documento'] = ''
                janela['status'].update('')
                driver.close()
        else:
            sg.popup_error('Digite algum documento!');
            janela['status'].update('')